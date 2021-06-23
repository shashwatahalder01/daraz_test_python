import openpyxl

def searchitems():
    
    productlist = []
    book = openpyxl.load_workbook('D:\\TestAutomation\\daraz_test_python\\data\\products.xlsx')
    sheet = book.active
    m_row = sheet.max_row

    for i in range(2, m_row + 1):
        cell = sheet.cell(row=i, column=1)
        productlist.append(cell.value)

    # print(productlist)

    return productlist
